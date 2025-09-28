# analytics.py — Assignment #2 (чистые визуалы, перцентильные фильтры, интерактив)

import os
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

from config import SQLALCHEMY_DATABASE_URL

CHARTS_DIR = "charts"
EXPORTS_DIR = "exports"
os.makedirs(CHARTS_DIR, exist_ok=True)
os.makedirs(EXPORTS_DIR, exist_ok=True)

engine = create_engine(SQLALCHEMY_DATABASE_URL, future=True)

# ---------- helpers ----------
def read_query_from_file(name: str, path: str = "queries.sql") -> str:
    with open(path, "r", encoding="utf-8") as f:
        sql_text = f.read()
    blocks = sql_text.split("\n----------------------------------------------------------------")
    for block in blocks:
        if f"name: {name}" in block:
            lines = [ln for ln in block.splitlines() if not ln.strip().startswith("--")]
            cleaned = "\n".join(lines).strip()
            return cleaned if cleaned.endswith(";") else cleaned + ";"
    raise ValueError(f"Query named '{name}' not found in {path}")

def fetch_df(query_name: str, parse_dates=None) -> pd.DataFrame:
    sql = read_query_from_file(query_name)
    df = pd.read_sql_query(text(sql), engine, parse_dates=parse_dates or [])
    return df.convert_dtypes()

def save_fig(fig, filename):
    path = os.path.join(CHARTS_DIR, filename)
    fig.savefig(path, bbox_inches="tight", dpi=150)
    plt.close(fig)
    return path

def report(df: pd.DataFrame, kind: str, meaning: str):
    print(f"{len(df)} rows → {kind}: {meaning}")

def cap_by_percentile(s, p=0.99):
    import numpy as np
    s_num = pd.to_numeric(s, errors="coerce")
    thr = np.nanquantile(s_num, p) if len(s_num) else None
    return s_num.clip(upper=thr) if thr is not None else s_num

# ---------- charts ----------
def chart_pie():
    df = fetch_df("pie_revenue_by_category")
    fig, ax = plt.subplots(figsize=(8, 8))
    ax.pie(df["revenue"], labels=df["category"], autopct="%1.1f%%", pctdistance=0.8)
    ax.set_title("Revenue share by product category (delivered, top-10+Other)")
    path = save_fig(fig, "01_pie_revenue_by_category.png")
    report(df, "Pie", "Delivered revenue share by category")
    return ("Pie", path, df)

def chart_bar():
    df = fetch_df("bar_top_sellers_by_revenue")
    # Сделаем подпись продавца покороче: город + хвост ID
    labels = df["seller_id"].astype(str).str.slice(0, 6) + "…" + df["seller_id"].astype(str).str[-4:]
    fig, ax = plt.subplots(figsize=(11, 7))
    ax.bar(labels, df["revenue"])
    ax.set_title("Top 10 sellers by delivered revenue")
    ax.set_xlabel("Seller (ID abridged)")
    ax.set_ylabel("Revenue")
    plt.xticks(rotation=30, ha="right")
    path = save_fig(fig, "02_bar_top_sellers_by_revenue.png")
    report(df, "Bar", "Top sellers by revenue")
    return ("Bar", path, df)

def chart_barh():
    df = fetch_df("barh_avg_review_by_category")
    df = df.sort_values(["avg_score", "n_reviews"], ascending=[False, False])
    fig, ax = plt.subplots(figsize=(10, 12))
    ax.barh(df["category"], df["avg_score"])
    ax.invert_yaxis()
    for i, (score, n) in enumerate(zip(df["avg_score"], df["n_reviews"])):
        ax.text(float(score), i, f"  {float(score):.2f} ({int(n)})", va="center", fontsize=8)
    ax.set_xlabel("Average review score")
    ax.set_ylabel("Category")
    ax.set_title("Average review by category (≥50 reviews), top-20")
    path = save_fig(fig, "03_barh_avg_review_by_category.png")
    report(df, "BarH", "Avg review score by category (top-20)")
    return ("BarH", path, df)

def chart_line():
    df = fetch_df("line_daily_revenue", parse_dates=["day"])
    df["revenue_capped"] = cap_by_percentile(df["revenue"], 0.99)
    fig, ax = plt.subplots(figsize=(11, 6))
    ax.plot(df["day"], df["revenue_capped"], marker="o", linewidth=1)
    ax.set_title("Daily delivered revenue (99th percentile capped)")
    ax.set_xlabel("Day")
    ax.set_ylabel("Revenue")
    path = save_fig(fig, "04_line_daily_revenue.png")
    report(df, "Line", "Daily delivered revenue (capped)")
    return ("Line", path, df)

def chart_hist():
    df = fetch_df("hist_order_value")
    s = pd.to_numeric(df["order_value"], errors="coerce").dropna()
    s = cap_by_percentile(s, 0.99)
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.hist(s, bins=30)
    ax.set_title("Distribution of order value (delivered, ≤99th pct)")
    ax.set_xlabel("Order value")
    ax.set_ylabel("Count")
    path = save_fig(fig, "05_hist_order_value.png")
    report(df, "Histogram", "Order value distribution (≤99th pct)")
    return ("Hist", path, df)

def chart_scatter():
    df = fetch_df("scatter_price_vs_review")
    df["avg_price"] = cap_by_percentile(df["avg_price"], 0.99)
    fig, ax = plt.subplots(figsize=(9, 6))
    ax.scatter(df["avg_price"], df["avg_review"], alpha=0.7)
    ax.set_title("Avg item price vs. average review (product-level)")
    ax.set_xlabel("Average item price (≤99th pct)")
    ax.set_ylabel("Average review score")
    path = save_fig(fig, "06_scatter_price_vs_review.png")
    report(df, "Scatter", "Price vs. review score by product (≤99th pct)")
    return ("Scatter", path, df)

# ---------- plotly time slider ----------
def show_time_slider(auto_open_html: bool = True):
    df = fetch_df("timeslider_monthly_revenue_by_country")
    fig = px.bar(
        df, x="country", y="revenue", animation_frame="month",
        title="Monthly delivered revenue by country"
    )
    if auto_open_html:
        out_html = os.path.join(CHARTS_DIR, "timeslider_revenue_by_country.html")
        fig.write_html(out_html, auto_open=True)
    else:
        fig.show()

# ---------- excel export ----------
def export_to_excel(frames, filename):
    path = os.path.join(EXPORTS_DIR, filename)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in frames.items():
            df.to_excel(writer, sheet_name=name, index=False)

    wb = load_workbook(path)
    for name in frames.keys():
        ws = wb[name]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        max_row, max_col = ws.max_row, ws.max_column
        for col in range(1, max_col + 1):
            numeric_samples = 0
            for r in range(2, min(max_row, 40) + 1):
                v = ws.cell(row=r, column=col).value
                if isinstance(v, (int, float)):
                    numeric_samples += 1
            if numeric_samples >= 3:
                col_letter = ws.cell(row=1, column=col).column_letter
                rng = f"{col_letter}2:{col_letter}{max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(rng, rule)
    wb.save(path)
    total_rows = sum(len(df) for df in frames.values())
    print(f"Created file {os.path.basename(path)}, {len(frames)} sheets, {total_rows} rows")
    return path

# ---------- main ----------
def main():
    charts = [
        chart_pie(),
        chart_bar(),
        chart_barh(),
        chart_line(),
        chart_hist(),
        chart_scatter(),
    ]
    frames = {kind: df for (kind, _path, df) in charts}
    export_to_excel(frames, "analytics_export.xlsx")
    # Для интерактива во время защиты:
    # show_time_slider()

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("[!] Error:", e)
        print("Hint: Check DB URL in config.py and column names.")
